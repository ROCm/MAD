"""Collecting a report's markdown and CSVs into one .xlsx, so postprocessing can happen there.

The workbook is a convenience over the markdown and the CSVs, which are the primary artifacts. A
missing openpyxl is therefore reported and skipped, not raised: raising here once threw away a
report that was already written and stopped the remaining phases from being produced at all.
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

RE_MD_BOLD = re.compile(r"\*\*(.+?)\*\*")
RE_MD_CODE = re.compile(r"`([^`]+)`")
RE_MD_RULE = re.compile(r"^:?-{2,}:?$")


def _as_number(value: str):
    """Spreadsheet cells should hold numbers, not strings that look like numbers."""
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        return value


def _plain(text: str) -> str:
    """Emphasis markers mean nothing in a cell, so only the words they wrap survive."""
    return RE_MD_CODE.sub(r"\1", RE_MD_BOLD.sub(r"\1", text)).strip()


def _md_cells(line: str) -> list:
    return [_plain(cell) for cell in line.strip().strip("|").split("|")]


def _write_report_sheet(ws, report_lines: list) -> None:
    """Render the markdown report as cells: tables become grids, prose stays in column A.

    Column A is kept narrow enough for the table it starts, so prose rows overflow across the empty
    cells to their right instead of being clipped.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    widths: dict = {}
    row = 0
    i = 0
    while i < len(report_lines):
        line = report_lines[i]
        if line.startswith("|"):
            block = []
            while i < len(report_lines) and report_lines[i].startswith("|"):
                block.append(_md_cells(report_lines[i]))
                i += 1
            aligns = []
            if len(block) > 1 and all(RE_MD_RULE.match(cell) for cell in block[1]):
                aligns = ["right" if cell.endswith(":") else "left" for cell in block[1]]
                del block[1]
            for r, cells in enumerate(block):
                row += 1
                for col, value in enumerate(cells, start=1):
                    cell = ws.cell(row, col, value if r == 0 else _as_number(value))
                    if r == 0:
                        cell.font = Font(bold=True)
                    elif col <= len(aligns):
                        cell.alignment = Alignment(horizontal=aligns[col - 1])
                    widths[col] = max(widths.get(col, 0), len(value))
            if i < len(report_lines) and report_lines[i].strip():
                row += 1  # keeps the next paragraph off the last table row
            continue

        i += 1
        row += 1
        if not line.strip():
            continue
        level = len(line) - len(line.lstrip("#"))
        cell = ws.cell(row, 1, _plain(line.lstrip("#").lstrip(">")))
        if level:
            cell.font = Font(bold=True, size=14 if level == 1 else 12)
        elif line.startswith(">"):
            cell.font = Font(italic=True)

    ws.column_dimensions["A"].width = min(max(widths.get(1, 0) + 2, 24), 46)
    for col, width in widths.items():
        if col > 1:
            ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 8), 46)


def write_workbook(out_dir: Path, report_lines: list) -> Path | None:
    """One workbook per report: the text as cells, every CSV as a sortable table.

    The rank x rank sheet stays a plain grid and gets a colour scale instead, which is the heatmap:
    openpyxl conditional formatting avoids a matplotlib dependency and keeps the result editable.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        print(f"warning: openpyxl is not installed in {sys.executable}, so "
              f"{out_dir}/profile.xlsx was skipped; report.md and the CSVs still hold every "
              "number. `pip install openpyxl` and rerun to get the workbook.")
        return None

    wb = Workbook()
    wb.remove(wb.active)
    _write_report_sheet(wb.create_sheet("report"), report_lines)

    for csv_path in sorted(out_dir.glob("*.csv")):
        with csv_path.open() as fh:
            rows = list(csv.reader(fh))
        if not rows:
            continue

        name = csv_path.stem[:31]
        is_matrix = name == "rank_matrix"
        header, body = rows[0], rows[1:]
        ws = wb.create_sheet(name)
        ws.append(header)
        for row in body:
            values = [_as_number(v) for v in row]
            if is_matrix:
                # Zeros would flatten the colour scale, so "no connection" stays empty.
                values = [values[0]] + [v or None for v in values[1:]]
            ws.append(values)

        if is_matrix:
            ws.freeze_panes = "B2"
            for cell in ws[1]:
                cell.font = Font(bold=True)
            last = get_column_letter(ws.max_column)
            ws.conditional_formatting.add(
                f"B2:{last}{ws.max_row}",
                ColorScaleRule(start_type="min", start_color="FFF7FBFF",
                               end_type="max", end_color="FF2166AC"),
            )
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 6
            continue

        ws.freeze_panes = "A2"
        table = Table(displayName=name,
                      ref=f"A1:{get_column_letter(len(header))}{len(body) + 1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
        ws.add_table(table)
        for col, title in enumerate(header, start=1):
            width = max([len(title)] + [len(row[col - 1]) for row in body if col <= len(row)])
            # +4 leaves room for the filter button the table adds to every header cell.
            ws.column_dimensions[get_column_letter(col)].width = min(max(width + 4, 10), 46)

    path = out_dir / "profile.xlsx"
    wb.save(path)
    return path
